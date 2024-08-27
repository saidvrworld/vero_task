
import argparse
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from datetime import datetime

import os

# Parses arguments and returns File,keys and colored variable
def parser():
    # Argument parsing
    parser = argparse.ArgumentParser(description="Client script for processing vehicle data and generating an Excel file.")
    parser.add_argument('-k', '--keys', nargs='+', help="Specify additional columns to include (e.g., kurzname, info).")
    parser.add_argument('-c', '--colored', action='store_true', default=True, help="Enable row coloring based on 'hu' field.")
    parser.add_argument('csv_file', help="Path to the input CSV file containing vehicle information.")
    args = parser.parse_args()
    return args.csv_file,args.keys,args.colored

def send_csv_with_post(file_path):
    # Send CSV to the server via POST request
    try:
        response = requests.post('http://127.0.0.1:8000/server/upload/', files={'file': open(file_path, 'rb')})
        #print(response.text)
        response.raise_for_status()
        server_data = response.json()
        return server_data
    except requests.exceptions.RequestException as e:
        print(f"Error communicating with the server: {e}")
        return None

def filter_server_response_and_save_it(file_path,columns_to_keep):
    server_data = send_csv_with_post(file_path)

    # Process the server's response
    df = pd.DataFrame(server_data)

    df = df.sort_values(by='gruppe')
    # Filter columns based on input keys
    required_columns = ['rnr','hu','labelIds','colorCodes'] + (columns_to_keep if columns_to_keep else [])
    df = df[[col for col in required_columns if col in df.columns]]


    output_filename = f"vehicles_{datetime.now().date().isoformat()}.xlsx"

    df.to_excel(output_filename, index=False)
    
    return output_filename

def get_column_index(ws, column_name):
    # Search for the column index based on the header row (assumed to be row 1)
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column

def color_text(excel_file):
    # Load the workbook and select the active worksheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Automatically get the column indices for 'labelIds' and 'colorCode'
    label_ids_col = get_column_index(ws, "labelIds")
    color_code_col = get_column_index(ws, "colorCodes")
    print(label_ids_col,color_code_col)
    if not label_ids_col or not color_code_col:
        print("Could not find 'labelIds' or 'colorCode' columns.")
        return

    # Iterate through each row in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
        label_id = row[label_ids_col - 1].value
        color_code = row[color_code_col - 1].value
        
        # Check if the labelIds and colorCode columns are not empty
        if label_id and color_code:
            # Apply the font color to all cells in this row
            for cell in row:
                cell.font = Font(color=color_code)

    # Save the changes
    wb.save(excel_file)
    print(f"Text colored in rows where labelIds and colorCode are both present.")

def calculate_month_difference(start_date, end_date):
    return (end_date.year - start_date.year) * 12 + end_date.month - start_date.month

def apply_row_color(ws, row, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in row:
        cell.fill = fill

def color_rows(excel_file):
    # Load the workbook and select the active worksheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Automatically get the column index for 'hu'
    hu_col = get_column_index(ws, "hu")

    if not hu_col:
        print("Could not find 'hu' column.")
        return

    # Define the color codes
    green = "007500"
    orange = "FFA500"
    red = "b30000"

    # Current date for comparison
    today = datetime.today()

    # Iterate through each row in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
        hu_date = row[hu_col - 1].value

        # Check if the hu column is not empty and is a valid date
        
        hu_date = datetime.strptime(hu_date, '%Y-%m-%d')
        if isinstance(hu_date, datetime):
            # Calculate the difference in months
            diff_in_months = calculate_month_difference(hu_date, today)
           
            # Determine the row color based on the date difference
            if diff_in_months <= 3:
                row_color = green
            elif diff_in_months <= 12:
                row_color = orange
            else:
                row_color = red

            # Apply the row color
            apply_row_color(ws, row, row_color)

    # Save the changes
    wb.save(excel_file)
    print(f"Rows colored based on the 'hu' column date.")


#########################################################################

local_file,passed_keys,colored =  parser()

output_filename = filter_server_response_and_save_it(local_file,passed_keys)

if("labelIds" in passed_keys):
    color_text(output_filename)

if colored:
    color_rows(output_filename)